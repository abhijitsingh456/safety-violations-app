from django.shortcuts import render
from django.contrib.auth.models import Group, User
from rest_framework import permissions, viewsets
from rest_framework.decorators import api_view
from .models import *
from .serializers import *
from rest_framework.response import Response
import openpyxl
import requests
from django.db.models import Q
from datetime import datetime, timedelta
from django.utils.timezone import make_aware

def index(request):
    return render(request, "safety_violation_app/speed_violators.html")

def search_speed_violators(request):
    return render(request, "safety_violation_app/search_speed_violators.html")

@api_view(['GET'])
def search_speed_violations(request, staffNo, vehicleNo, start_date, end_date, speed):
    if request.method=="GET":
        start_date = datetime.strptime(str(start_date)+" 00:00:01", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
        end_date = datetime.strptime(str(end_date)+" 23:59:59", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
        if (vehicleNo=="any" and staffNo!="any"):
            query_result = SpeedViolations.objects.filter(employee__staff_no=staffNo, date__gte=start_date, date__lte=end_date, speed__gte=speed).all()      
        elif (staffNo=="any" and vehicleNo!="any"):
            query_result = SpeedViolations.objects.filter(plate_text=vehicleNo, date__gte=start_date, date__lte=end_date, speed__gte=speed).all()
        elif(vehicleNo=="any" and staffNo=="any"):
            query_result = SpeedViolations.objects.filter(date__gte=start_date, date__lte=end_date, speed__gte=speed).all()
        serializer = ViolationSerializer(query_result, many=True)
        return Response(serializer.data)


@api_view(['GET','POST'])
def get_employees(request):
    if request.method=="GET":
        employees = Employee.objects.all()
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)
    if request.method=='POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error':'No File Uploaded'},status=400)
        if not uploaded_file.name.endswith(('.xlsx','xls','.XLSX','.XLS')):
            return Response({'error':'Invaid File Type'},status=400)
        
        try:
            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active

            data = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if (i==0):
                    header = list(row)
                    req_headers = ['STAFF NUMBER','FIRST NAME', 'MIDDLE NAME', 'LAST NAME', 'department name', 'Designation', 'PHONE NUMBER', 'EMAIL ADDRESS']
                    for h in req_headers:
                        if h not in header:
                            return Response({'error':'Invaid Column Headers in File '},status=400)
                    Employee.objects.all().delete()
                    continue

                employee_row = dict(zip(header,row))
                e = Employee(staff_no=str(employee_row['STAFF NUMBER']), 
                            name=employee_row['FIRST NAME']+" "+employee_row['MIDDLE NAME']+" "+employee_row['LAST NAME'], 
                            department=employee_row['department name'], 
                            designation=employee_row['Designation'], 
                            phone_no=str(employee_row['PHONE NUMBER']),
                            email_address=employee_row['EMAIL ADDRESS'])
                e.save()
            return Response({'message':'File uploaded successfully'},status=200)
        except Exception as e:
            return Response({'error':str(e)},status=500)

@api_view(['GET','POST'])
def get_vehicle_details(request):
    if request.method=="GET":
        vehicles = Vehicles.objects.all()
        serializer = VehicleSerializer(vehicles, many=True)
        return Response(serializer.data)
    if request.method=='POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error':'No File Uploaded'},status=400)
        if not uploaded_file.name.endswith(('.xlsx','xls','.XLSX','.XLS')):
            return Response({'error':'Invaid File Type'},status=400)
        
        try:
            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active

            data = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if (i==0):
                    header = list(row)
                    req_headers = ['STNO','NAME','DEPARTMENT','DESIGNATION','FIRST VEHICLE','SECOND VEHICLE']
                    for h in req_headers:
                        if h not in header:
                            return Response({'error':'Invaid Column Headers in File '},status=400)
                    Vehicles.objects.all().delete()
                    continue
                vehicle_row = dict(zip(header,row))

                v = Vehicles(staff_no=str(vehicle_row['STNO']), 
                            name=vehicle_row['NAME'], 
                            department=vehicle_row['DEPARTMENT'], 
                            designation=vehicle_row['DESIGNATION'], 
                            first_vehicle=vehicle_row['FIRST VEHICLE'].split("(")[0],
                            second_vehicle=vehicle_row['SECOND VEHICLE'].split("(")[0])
                v.save()
            return Response({'message':'File uploaded successfully'},status=200)
        except Exception as e:
            return Response({'error':str(e)},status=500)

@api_view(['GET','POST'])
def update_speed_violations(request): #used to update the database with violations of the previous day
        start_date = (datetime.strptime(str((datetime.now() - timedelta(1)).strftime('%Y-%m-%d'))+" 00:00:00", '%Y-%m-%d %H:%M:%S')).strftime('%Y-%m-%d %H:%M:%S')
        end_date = (datetime.strptime(str((datetime.now() - timedelta(1)).strftime('%Y-%m-%d'))+" 23:59:00", '%Y-%m-%d %H:%M:%S')).strftime('%Y-%m-%d %H:%M:%S')
        url = 'https://api.streetsoncloud.com/v4/tickets-list'
        headers = {'Content-Type': 'application/json',
                        'Range': 'items=0-10',
                        'x-api-key':'0YjoQrlmP87aseagqbw2i75weVuTONplavJn4UFu',
                        'X-Requested-With':'XMLHttpRequest'}
        final_result = []
        for i in range(-4,10):
                data = {"dateFrom":start_date,
                        "dateTo": end_date,
                        "exportStatus": None,
                        "status":i,
                        "accountId":5574,
                        "units": 1,
                        "optional": True}
                results = requests.post(url, json=data, headers = headers)
                if (results.status_code==200):
                        final_result+=results.json()['tickets']
                else:
                        return Response({'error':'streetsoncloud API error'},status=results.status_code) 
        for violation in final_result:
                vehicle = Vehicles.objects.filter(
                Q(first_vehicle=violation['plateText']) | Q(second_vehicle=violation['plateText'])
                ).first()
                if vehicle:
                        try:
                                employee = Employee.objects.get(staff_no=vehicle.staff_no)
                                speed_violation = SpeedViolations(
                                employee=employee,
                                date=violation['dateTimeLocal'],
                                location=violation['location'],
                                plate_text=violation['plateText'],
                                speed=int(float((violation['speedObserved'])) * 1.60934)  # mph to km/h
                                )
                                speed_violation.save()
                        except Employee.DoesNotExist:
                                print(f"Employee with staff_no {vehicle.staff_no} does not exist.")
                else:
                        print(f"No vehicle match found for plateText: {violation['plateText']}")
        return Response({'message':'Databse updated successfully'},status=200)

@api_view(['GET','POST'])
def get_no_of_speed_violations(request, start_date, end_date, no_of_violations, speed):
    start_date = datetime.strptime(str(start_date)+" 00:00:01", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
    end_date = datetime.strptime(str(end_date)+" 23:59:59", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
    query_result = SpeedViolations.objects.filter(date__gte=start_date, date__lte=end_date, speed__gte=speed)
    result = []
    staff_nos = []
    for q in query_result:
        if (q.employee.staff_no in staff_nos):
            pass
        else:
            staff_nos.append(q.employee.staff_no)
            count = query_result.filter(employee__staff_no=q.employee.staff_no).count()
            if (count==no_of_violations  or (no_of_violations==6 and count>=6) or (no_of_violations==0 and count>=0)):
                result.append({"staff_no":q.employee.staff_no,
                            "name":q.employee.name,
                            "department":q.employee.department,
                            "designation":q.employee.designation,
                            "no_of_violations":count})
    serializer = NoOfSpeedViolationSerializer(result, many=True)
    return Response(serializer.data)
